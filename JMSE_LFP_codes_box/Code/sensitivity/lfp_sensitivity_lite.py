"""
lfp_sensitivity_lite.py
=======================
Lightweight LFP sensitivity study that depends ONLY on the five indicator
scripts (pci / ci / cei / ecpi / mre) -- it does NOT re-derive the bill-of-
materials from GREET (no build_lfp_inputs.py / build_lfp_mre.py / lfp_greet_data.py).

Idea: the GREET-derived base case is invariant across scenarios; each scenario
is just a scalar knob applied to a handful of cells. So instead of rebuilding
every input CSV from GREET nine times, we cache the base-case CSVs once and
transform only the affected cells, then run the indicator scripts.

Studies (identical design to build_lfp_sensitivity.py, the reference of record):
  S1  recycling energy   x {+20, 0, -20}%   -> CI, E
  S2  recovery efficiency x {+5, 0, -5}%    -> PCI, CI, CEI, M
  S3  collection efficiency {70, 90, 100}%  -> PCI, ECPI, CEI, M, R, E

The few cell-level transforms below were derived from build_lfp_inputs.py /
build_lfp_mre.py and are proven faithful by `--validate-inputs`, which rebuilds
each scenario's CSVs the orchestrator way and diffs them cell-for-cell.

A study run leaves the working tree's cumulative_results_indicators.csv and the
indicator PNGs in last-scenario state; main() finishes by regenerating the base
case so the tree is left clean. Run `--restore-base` to do that on its own.

Usage:
  python lfp_sensitivity_lite.py                   # S1/S2/S3 *_lite tables, plots, Raw_data (then restores base)
  python lfp_sensitivity_lite.py --validate-inputs # deterministic check vs the orchestrator (source layout only)
  python lfp_sensitivity_lite.py --restore-base    # restore base CSVs + rebuild base cumulative & figures
"""
import os, sys, csv, io, subprocess, shutil
import numpy as np
import matplotlib; matplotlib.use('Agg')
import matplotlib.pyplot as plt
try:
    import openpyxl
except ImportError:
    openpyxl = None

HERE = os.path.dirname(os.path.abspath(__file__))
PY   = sys.executable
ENV  = dict(os.environ, MPLBACKEND='Agg', PYTHONIOENCODING='utf-8')
COLL_BASE = 0.90
# (ECPI now computes all GHG inside ecpi_indicator.py from raw inputs; the sensitivity
#  transform only nudges PARAM_collection / F_r, so no recycled-GHG constant is needed here.)

# ---- locate indicator dirs / csvs / scripts (works in source LFP/ and in box Code/sensitivity/) ----
IND = {  # key: (dir-name candidates, csv filename, indicator script)
    'pci':  (['pci','PCI'],  'pci_material_parameters.csv',   'pci_indicator_v2.py'),
    'ci':   (['ci','CI'],    'ci_parameters.csv',             'ci_indicator.py'),
    'cei':  (['cei','CEI'],  'cei_parameters.csv',            'cei_indicator.py'),
    'ecpi': (['ecpi','ECPI'],'ECPI_indicator_alpha_sum.csv',  'ecpi_indicator.py'),
    'mre':  (['mre','MRE'],  'mre_indicator.csv',             'mre_indicator.py'),
}
def _find_dir(cands):
    for base in (HERE, os.path.join(HERE, '..')):
        for c in cands:
            d = os.path.join(base, c)
            if os.path.isdir(d):
                return os.path.abspath(d)
    raise FileNotFoundError(f"indicator dir not found among {cands}")
DIRS   = {k: _find_dir(v[0]) for k, v in IND.items()}
CSVP   = {k: os.path.join(DIRS[k], IND[k][1]) for k in IND}
SCRIPT = {k: IND[k][2] for k in IND}
CUM    = os.path.join(DIRS['pci'], '..', 'cumulative_results_indicators.csv')

# ---- base-case snapshot (taken at import; the working tree must be at the base case) ----
BASE_TEXT = {k: open(CSVP[k], 'r', encoding='utf-8').read() for k in IND}

def _rows(text):
    return list(csv.reader(io.StringIO(text)))
def _text(rows):
    buf = io.StringIO(); csv.writer(buf, lineterminator='\n').writerows(rows); return buf.getvalue()
def _write(path, text):
    try:
        with open(path, 'w', encoding='utf-8', newline='') as f: f.write(text)
    except PermissionError:
        raise SystemExit(f"\nERROR: cannot write {os.path.relpath(path, HERE)} -- the file is locked "
                         f"(open in Excel?). Close it and re-run.\n")

def restore_base():
    for k in IND:
        _write(CSVP[k], BASE_TEXT[k])

def _isnum(s):
    try: float(s); return True
    except (TypeError, ValueError): return False

# ======================================================================== transforms
def t_ci(text, recyc=1.0, recov=1.0):
    rows = _rows(text); out = []
    for r in rows:
        r = list(r); name = r[0] if r else ''
        # energy_recycling (col 4) x recyc  for steel / aluminum / cell_total
        if recyc != 1.0 and name in ('steel', 'aluminum', 'cell_total') and len(r) > 4 and _isnum(r[4]):
            r[4] = repr(float(r[4]) * recyc)
        # recovery_rate (col 2) -> min(1, base x recov)  (steel, aluminum, cell_material_*)
        if recov != 1.0 and len(r) > 2 and _isnum(r[2]) and name not in ('cell_total',):
            r[2] = repr(min(1.0, float(r[2]) * recov))
        out.append(r)
    return _text(out)

def t_pci(text, recov=1.0, coll=None):
    rows = _rows(text); out = []
    for r in rows:
        r = list(r); name = r[0] if r else ''
        if name not in ('material', 'TOTAL_BATTERY_MASS') and len(r) > 5 and _isnum(r[1]):
            if recov != 1.0 and _isnum(r[3]): r[3] = repr(min(1.0, float(r[3]) * recov))  # F_r
            if coll is not None:              r[5] = repr(coll)                           # C_r
        out.append(r)
    return _text(out)

def t_cei(text, recov=1.0, coll=None):
    rows = _rows(text); out = [rows[0]]
    c = COLL_BASE if coll is None else coll
    for r in rows[1:]:
        r = list(r)
        mass = float(r[1]); rec_base = float(r[2])
        recov_base = rec_base / (mass * COLL_BASE) if mass else 0.0
        r[2] = repr(mass * min(1.0, recov_base * recov) * c)   # mass_recycled = mass x recov_eff x collection
        out.append(r)
    return _text(out)

def t_ecpi(text, recov=1.0, coll=None):
    """ECPI raw-input schema (revamp 2026-06-21). All GHG/alpha/beta math is now in
    ecpi_indicator.py, so a sweep only touches the raw knobs:
      collection -> PARAM_collection value (col 1)
      recovery   -> each material's F_r (col 4), x recov, capped at 1.0
    Mirrors build_lfp_inputs.py (LFP_COLLECTION / LFP_RECOVERY_MULT); proven by
    --validate-inputs. Recovery now moves ECPI (alpha_out, and beta via the steel
    direct rate), so ECPI is reported in S2 as well as S3."""
    rows = _rows(text); out = []
    for r in rows:
        r = list(r); name = r[0] if r else ''
        if name == 'PARAM_collection':
            if coll is not None:
                r[1] = repr(coll)                                    # PARAM value sits in col 1
        elif not name.startswith('PARAM_') and name != 'material':
            if recov != 1.0 and len(r) > 4 and _isnum(r[4]):
                r[4] = repr(min(1.0, float(r[4]) * recov))           # F_r x recov, capped at 1
        out.append(r)
    return _text(out)

def t_mre(text, recyc=1.0, recov=1.0, coll=None):
    lines = text.splitlines()
    # locate base m_re_total / m_vir_total for the collection rescale
    def _pv(nm):
        for ln in lines:
            p = ln.split(',')
            if p and p[0] == nm and len(p) > 1 and _isnum(p[1]): return float(p[1])
        return None
    mre_b, mvir_b = _pv('m_re_total'), _pv('m_vir_total')
    cs = None if coll is None else coll / COLL_BASE
    out = []
    for ln in lines:
        p = ln.split(','); nm = p[0] if p else ''
        if nm == 'Material_Data' and recov != 1.0 and len(p) > 9 and _isnum(p[9]):
            p[9] = repr(min(1.0, float(p[9]) * recov))                       # eta_hydro
        elif nm in ('E_recovery', 'E_pre_treatment') and recyc != 1.0 and _isnum(p[1]):
            p[1] = repr(float(p[1]) * recyc)                                 # recycling energy (S1)
        elif nm in ('m_remanufactured', 'm_recycled_at_80') and cs is not None and _isnum(p[1]):
            p[1] = repr(float(p[1]) * cs)                                    # collection-sensitive recovery mass
        elif nm == 'm_re_total' and cs is not None:
            p[1] = repr(mre_b * cs)
        elif nm == 'm_vir_total' and cs is not None:
            p[1] = repr((mre_b + mvir_b) - mre_b * cs)
        out.append(','.join(p))
    return '\n'.join(out) + '\n'

# scenario -> dict {ind_key: perturbed_text}; only the indicators a study reports are perturbed/run
def make_inputs(study, knob):
    if study == 'S1':   # recycling energy
        m = knob
        return {'ci': t_ci(BASE_TEXT['ci'], recyc=m), 'mre': t_mre(BASE_TEXT['mre'], recyc=m)}
    if study == 'S2':   # recovery efficiency
        m = knob
        return {'pci': t_pci(BASE_TEXT['pci'], recov=m), 'ci': t_ci(BASE_TEXT['ci'], recov=m),
                'cei': t_cei(BASE_TEXT['cei'], recov=m), 'ecpi': t_ecpi(BASE_TEXT['ecpi'], recov=m),
                'mre': t_mre(BASE_TEXT['mre'], recov=m)}
    if study == 'S3':   # collection efficiency
        c = knob
        return {'pci': t_pci(BASE_TEXT['pci'], coll=c), 'ecpi': t_ecpi(BASE_TEXT['ecpi'], coll=c),
                'cei': t_cei(BASE_TEXT['cei'], coll=c), 'mre': t_mre(BASE_TEXT['mre'], coll=c)}
    raise ValueError(study)

STUDIES = {
    'S1': dict(col='Recycling Energy Change (%)', inds=['CI', 'E'],
               scen=[(20, 1.2), (0, 1.0), (-20, 0.8)], title='Impact of Recycling Energy Change'),
    'S2': dict(col='Recovery Efficiency Change (%)', inds=['PCI', 'CI', 'CEI', 'ECPI', 'M'],
               scen=[(5, 1.05), (0, 1.0), (-5, 0.95)], title='Impact of Recovery Efficiency Change'),
    'S3': dict(col='Collection Efficiency (%)', inds=['PCI', 'ECPI', 'CEI', 'M', 'R', 'E'],
               scen=[(70, 0.70), (90, 0.90), (100, 1.00)], title='Impact of Collection Efficiency'),
}
PRODUCER = {'PCI': 'pci', 'CI': 'ci', 'CEI': 'cei', 'ECPI': 'ecpi', 'M': 'mre', 'R': 'mre', 'E': 'mre'}

def run_indicator(k):
    r = subprocess.run([PY, '-X', 'utf8', SCRIPT[k]], cwd=DIRS[k], env=ENV, capture_output=True, text=True)
    if r.returncode:
        print('ERR', k); print(r.stderr[-1200:]); raise SystemExit(1)

def read_cum():
    out = {}
    with open(CUM) as f:
        for row in csv.DictReader(f):
            out[row['Indicator']] = (float(row['Mean']), float(row['CI_Lower']), float(row['CI_Upper']))
    return out

def run_scenario(study, knob):
    inputs = make_inputs(study, knob)
    restore_base()
    for k, txt in inputs.items():
        _write(CSVP[k], txt)
    for k in inputs:                      # run only the perturbed indicators
        run_indicator(k)
    data = read_cum()
    restore_base()
    return data

def regenerate_base():
    """Restore the base-case input CSVs and rebuild cumulative_results_indicators.csv
    plus the base-case indicator + compiled figures -- undoes the last-scenario state a
    study run leaves behind. Assumes the cached base snapshot is the true base case."""
    print("Regenerating base case (restore CSVs -> run 5 indicators -> compiled)...")
    restore_base()
    for k in ('pci', 'ci', 'cei', 'ecpi', 'mre'):
        print(f"  base run {k}/{SCRIPT[k]}")
        run_indicator(k)
    comp_dir = os.path.dirname(os.path.abspath(CUM))
    if os.path.exists(os.path.join(comp_dir, 'compiled_results_indicator.py')):
        r = subprocess.run([PY, '-X', 'utf8', 'compiled_results_indicator.py'],
                           cwd=comp_dir, env=ENV, capture_output=True, text=True)
        if r.returncode:
            print('  compiled WARN:', r.stderr[-300:])
        else:
            src = os.path.join(comp_dir, 'compiled_indicators_comparison.png')   # script's fixed name
            dst = os.path.join(comp_dir, 'compiled_indicators_comparison_lfp.png')  # deliverable name
            if os.path.exists(src) and os.path.exists(dst):
                shutil.copyfile(src, dst)
    print("Base case restored: cumulative + figures rebuilt.")

# ======================================================================== outputs
def plot_study(rows, col, inds, title, out_png):
    x = np.arange(len(inds)); w = 0.25; colors = ['#5b9bd5', '#ed7d31', '#a5a5a5']
    fig, ax = plt.subplots(figsize=(14, 7))
    for i, (sval, data) in enumerate(rows):
        means = [data[k][0] for k in inds]
        lo = [max(0.0, data[k][0]-data[k][1]) for k in inds]; hi = [max(0.0, data[k][2]-data[k][0]) for k in inds]
        ax.bar(x+(i-1)*w, means, w, color=colors[i], edgecolor='black', linewidth=0.8, label=f'{sval}%')
        ax.errorbar(x+(i-1)*w, means, yerr=[lo, hi], fmt='none', ecolor='black', elinewidth=1.3, capsize=4, capthick=1.3)
    ax.set_xticks(x); ax.set_xticklabels(inds, fontsize=13, fontweight='bold')
    ax.set_ylabel('Indicator Value', fontsize=13, fontweight='bold')
    ax.set_xlabel('Circularity Indicators', fontsize=13, fontweight='bold')
    ax.set_title(f'Sensitivity Analysis (lite): {title}\n(LFP, with 95% Confidence Intervals)', fontsize=14, fontweight='bold')
    ax.legend(title=col.replace(' (%)', ''), fontsize=11); ax.grid(axis='y', alpha=0.3, ls='--'); ax.set_axisbelow(True)
    ax.set_ylim(0, 0.78); plt.tight_layout(); plt.savefig(out_png, dpi=300, bbox_inches='tight'); plt.close()

def write_xlsx(path, col, inds, rows):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Sheet1'
    hdr = [col]
    for ind in inds: hdr += [f'{ind} Mean', f'{ind} CI Lower', f'{ind} CI Upper']
    ws.append(hdr)
    for sval, data in rows:
        r = [sval]
        for ind in inds:
            m, lo, hi = data[ind]; r += [round(m, 4), round(lo, 4), round(hi, 4)]
        ws.append(r)
    wb.save(path)

def _base_is_valid():
    """Guard against snapshotting a poisoned tree: a study run perturbs the input CSVs in
    place, so if a PREVIOUS run crashed mid-scenario the on-disk CSVs (and thus BASE_TEXT,
    taken at import) are NOT the base case. Check the PCI collection column == COLL_BASE."""
    for r in _rows(BASE_TEXT['pci']):
        if r and r[0] not in ('material', 'TOTAL_BATTERY_MASS') and len(r) > 5 and _isnum(r[5]):
            if abs(float(r[5]) - COLL_BASE) > 1e-6:
                return False
    return True

def main():
    if not _base_is_valid():
        raise SystemExit(
            "\nERROR: the input CSVs are NOT at the base case (a previous run likely crashed "
            "mid-scenario, e.g. on a locked file). Re-copy the pristine base parameter CSVs into\n"
            "  pci/ ci/ cei/ ecpi/ mre/  and retry. (PCI collection column != 0.90 detected.)\n")
    raw = openpyxl.Workbook() if openpyxl else None
    if raw: raw.remove(raw.active)
    all_rows = {}
    try:
      for sid, cfg in STUDIES.items():
        print(f"\n=== {sid} (lite): {cfg['title']} ===")
        rows = []
        for sval, knob in cfg['scen']:
            data = run_scenario(sid, knob)
            rows.append((sval, data))
            print(f"  {sval:>4}: " + " ".join(f"{k}={data[k][0]:.3f}" for k in cfg['inds']))
        all_rows[sid] = rows
        if openpyxl:
            write_xlsx(os.path.join(HERE, f'{sid}_sensitivity_study_lfp.xlsx'), cfg['col'], cfg['inds'], rows)
            ws = raw.create_sheet(sid); hdr = [cfg['col']]
            for ind in cfg['inds']: hdr += [f'{ind} Mean', f'{ind} CI Lower', f'{ind} CI Upper']
            ws.append(hdr)
            for sval, data in rows:
                r = [sval]
                for ind in cfg['inds']:
                    m, lo, hi = data[ind]; r += [round(m, 6), round(lo, 6), round(hi, 6)]
                ws.append(r)
        plot_study(rows, cfg['col'], cfg['inds'], cfg['title'],
                   os.path.join(HERE, f'{sid}_sensitivity_scenarios_comparison_lfp.png'))
      if raw: raw.save(os.path.join(HERE, 'Raw_data_sensitivity_lfp.xlsx'))
      print("\nWrote S1/S2/S3_sensitivity_study_lfp.xlsx, comparison PNGs, Raw_data_sensitivity_lfp.xlsx")
    finally:
        # ALWAYS leave the tree at a clean base case, even if a scenario crashed (e.g. on a
        # locked file) -- prevents a perturbed tree from poisoning the next run's base snapshot.
        try:
            regenerate_base()
        except SystemExit:
            pass   # restore_base() inside regenerate_base already reverted the input CSVs

# ======================================================================== input-level validation
def validate_inputs():
    """For every scenario, rebuild the CSVs the orchestrator way (build_lfp_inputs /
    build_lfp_mre with env knobs) and diff them cell-for-cell against the lite transforms.
    Deterministic (no Monte Carlo)."""
    def orch(env_over):
        e = dict(ENV, LFP_COLLECTION='0.90', LFP_ROUTE='hydro'); e.update(env_over)
        for s in ('build_lfp_inputs.py', 'build_lfp_mre.py'):
            r = subprocess.run([PY, '-X', 'utf8', s], cwd=HERE, env=e, capture_output=True, text=True)
            if r.returncode: print('ORCH ERR', s); print(r.stderr[-800:]); raise SystemExit(1)
        return {k: open(CSVP[k], encoding='utf-8').read() for k in IND}

    def num_diff(a, b):
        ra, rb = _rows(a), _rows(b); worst = 0.0; where = ''
        for i, (rowa, rowb) in enumerate(zip(ra, rb)):
            for j, (ca, cb) in enumerate(zip(rowa, rowb)):
                if _isnum(ca) and _isnum(cb):
                    d = abs(float(ca) - float(cb))
                    if d > worst: worst, where = d, f'r{i}c{j} ({rowa[0]})'
        return worst, where

    SCN = {  # study: list of (label, lite_inputs_fn_args, orch_env)
        'S1': [(20, dict(recyc=1.2), {'LFP_RECYC_MULT': '1.2'}), (-20, dict(recyc=0.8), {'LFP_RECYC_MULT': '0.8'})],
        'S2': [(5, dict(recov=1.05), {'LFP_RECOVERY_MULT': '1.05'}), (-5, dict(recov=0.95), {'LFP_RECOVERY_MULT': '0.95'})],
        'S3': [(70, dict(coll=0.70), {'LFP_COLLECTION': '0.70'}), (100, dict(coll=1.00), {'LFP_COLLECTION': '1.00'})],
    }
    LITE = {  # which csvs each study perturbs (others identical to base by construction)
        'S1': lambda kw: {'ci': t_ci(BASE_TEXT['ci'], recyc=kw['recyc']), 'mre': t_mre(BASE_TEXT['mre'], recyc=kw['recyc'])},
        'S2': lambda kw: {'pci': t_pci(BASE_TEXT['pci'], recov=kw['recov']), 'ci': t_ci(BASE_TEXT['ci'], recov=kw['recov']),
                          'cei': t_cei(BASE_TEXT['cei'], recov=kw['recov']), 'ecpi': t_ecpi(BASE_TEXT['ecpi'], recov=kw['recov']),
                          'mre': t_mre(BASE_TEXT['mre'], recov=kw['recov'])},
        'S3': lambda kw: {'pci': t_pci(BASE_TEXT['pci'], coll=kw['coll']), 'ecpi': t_ecpi(BASE_TEXT['ecpi'], coll=kw['coll']),
                          'cei': t_cei(BASE_TEXT['cei'], coll=kw['coll']), 'mre': t_mre(BASE_TEXT['mre'], coll=kw['coll'])},
    }
    print("Input-level validation: lite transforms vs orchestrator (build_lfp_inputs/mre)\n" + "-"*70)
    worst_overall = 0.0
    try:
        for sid, scns in SCN.items():
            for label, kw, env_over in scns:
                ref = orch(env_over)                      # orchestrator-built CSVs (overwrites working tree)
                lite = LITE[sid](kw)                      # lite-built CSVs (from cached base)
                line = f"{sid} {label:>4}%: "
                for k in lite:
                    d, where = num_diff(lite[k], ref[k])
                    worst_overall = max(worst_overall, d)
                    line += f"{k}:{d:.2e} "
                print(line)
    finally:
        restore_base()
    print("-"*70)
    tol = 1e-3
    print(f"max abs cell diff = {worst_overall:.3e}   ->   {'PASS' if worst_overall < tol else 'FAIL'} (tol {tol})")

if __name__ == '__main__':
    if '--validate-inputs' in sys.argv:
        validate_inputs()
    elif '--restore-base' in sys.argv:
        regenerate_base()
    else:
        main()
